import sendmail, task, atlogger
import pyodbc
import os, datetime, sys
from openpyxl import Workbook
from config import config
from openpyxl.cell import get_column_letter
from openpyxl.worksheet.page import PageMargins

class task:
    def __init__(self, t, logger):
        self.taskinfo = t
        self.logger = logger

    def resultIter(self, cursor, size=300):
        while True:
            results = cursor.fetchmany(size)
            if not results: break
            for result in results:
                yield result


    def sendFile(self, fname):
        login = self.taskinfo.get('sender'), self.taskinfo.get('mailpass')
        cc = self.taskinfo.get('cc')
        bcc = self.taskinfo.get('bcc')
        to = self.taskinfo.get('receiver')
        sender = self.taskinfo.get('sender')
        subject = self.taskinfo.get('mailtitle')
        text = self.taskinfo.get('mailtext')
        server = self.taskinfo.get('mailserver')
        sendmail.sendmail(login, sender, to, cc, bcc, subject, text, fname, server)  

    def work(self):
        try:
            sql = None
            curdir = os.path.dirname(os.path.realpath(__file__)) 
            con_str = self.taskinfo.get('dbconn')
            con = pyodbc.connect(con_str) 
            cursor = con.cursor()
            wb = Workbook()
            templatesheet = self.taskinfo.get('templatesheet')
            for idx, sh in enumerate(self.taskinfo.get('sheets')):
                sheet = {}
                sheet.update(templatesheet)
                sheet.update(sh)
                sqlfile = sheet.get('sql') 
                if os.sep not in sqlfile:
                    sqlfile = os.path.join(curdir,'sql') + os.sep + sqlfile
                with open(sqlfile) as fn:
                    sql = fn.read()
                ws = None
                if idx == 0:
                    ws = wb.worksheets[0]
                else: ws = wb.create_sheet()
                ws.title = sheet.get('name')
                #margin
                margins = sheet.get('margin')
                if margins:
                    ws.page_margins = PageMargins(margins[0], margins[1], margins[2], margins[3], margins[4], margins[5])
                #fitpage
                if sheet.get('fitpage','true').strip() == 'true':
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToHeight = 0
                    ws.page_setup.fitToWidth = 1
                #orientation
                ort = sheet.get('orientation','landscape')
                if ort in ['portrait','landscape']:
                    ws.page_setup.orientation = ort
                #gridline
                if sheet.get('printgrid','true').strip() == 'true':
                    ws.print_grid = True
                #rowtitle
                titlerow = sheet.get('titlerow')
                if titlerow is not None:
                    ws.add_print_title( sheet.get('titlerow'), rows_or_cols = 'rows')
                #paper size:
                #PAPERSIZE_A3, PAPERSIZE_A4, PAPERSIZE_A4_SMALL, PAPERSIZE_A5, PAPERSIZE_EXECUTIVE
                #PAPERSIZE_LEDGER, PAPERSIZE_LEGAL, PAPERSIZE_LETTER, PAPERSIZE_LETTER_SMALL, PAPERSIZE_TABLOID
                psize = getattr(ws, sheet.get('papersize','PAPERSIZE_A4').upper())
                if psize:
                    ws.page_setup.paperSize = psize
                header = sheet.get('header')
                if header and isinstance(header, list) and len(header)==3:
                    if header[0] and header[0].strip() != '': 
                        ws.header_footer.left_header.text = header[0]
                    if header[1] and header[1].strip() != '': 
                        ws.header_footer.center_header.text = header[1]
                    if header[2] and header[2].strip() != '': 
                        ws.header_footer.right_header.text = header[2]
                footer = sheet.get('footer')
                if footer and isinstance(footer, list) and len(footer)==3:
                    if footer[0] and footer[0].strip() != '':
                        ws.header_footer.left_footer.text = footer[0]
                    if footer[1] and footer[1].strip() != '':
                        ws.header_footer.center_footer.text = footer[1]
                    if footer[2] and footer[2].strip() != '':
                        ws.header_footer.right_footer.text = footer[2]
                cursor.execute(sql)
                columns = [column[0] for column in cursor.description]
                fmt = []
                for c in columns:
                    if '$' in c: fmt.append(1)
                    elif '%' in c: fmt.append(2)
                    else: fmt.append(0)
                # Make column header
                ws.append(columns)
                # format and column width
                column_widths = [0]*len(columns)
                for r in cursor:
                    ws.append(list(r))
                for row in ws.iter_rows():
                    for i, cell in enumerate(row):
                        if fmt[i] == 1 : cell.number_format = '$#,##0_-'
                        elif fmt[i] == 2 : cell.number_format =  '0.00%'
                        if column_widths[i] < len(str(cell.value)): column_widths[i] = len(str(cell.value))
                    
                for i, c in enumerate(column_widths):
                    #print(c)
                    ws.column_dimensions[get_column_letter(i+1)].width = c
                    
            con.close()
            # Generate excel file
            fname = os.path.join(curdir,'excel') + os.sep + self.taskinfo.get('excel') + '-' + datetime.datetime.now().strftime('%Y%m%d-%H%M%S.xlsx')
            wb.save(filename = fname )
            self.sendFile([fname])
        except Exception as e:
            self.logger.exception(e)

if __name__ == '__main__':
    atlogger.g_logger.addScreenMode()
    atlogger.g_logger.setDebug()
    c = config()
    if not c.loadCfg(): sys.exit()
    for t in c.tasksAll():
        tsk = task(t, atlogger.g_logger)
        tsk.work()
